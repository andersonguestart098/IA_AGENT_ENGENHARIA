# app/ingest/xlsx_extractor.py
from __future__ import annotations

import io
import re
import hashlib
from datetime import datetime, date
from typing import Any, Dict, List, Optional, Tuple

import openpyxl


def _sha256_bytes(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _norm_header(s: str) -> str:
    s = (s or "").strip()
    s = re.sub(r"\s+", " ", s)
    return s


def _infer_cell_type(v: Any) -> str:
    if v is None:
        return "null"
    if isinstance(v, bool):
        return "bool"
    if isinstance(v, int):
        return "int"
    if isinstance(v, float):
        return "float"
    if isinstance(v, (datetime, date)):
        return "date"
    return "str"


def _safe_str(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float):
        s = f"{v:.10f}".rstrip("0").rstrip(".")
        return s
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return str(v).strip()


def _detect_header_row(values: List[List[Any]], max_scan: int = 25) -> int:
    """
    Heurística simples e boa:
    - varre as primeiras linhas e escolhe a que tem mais "strings úteis" (colunas nomeadas)
    - evita linhas quase vazias
    Retorna índice 0-based.
    """
    best_i = 0
    best_score = -1

    scan = values[: max_scan or 1]
    for i, row in enumerate(scan):
        if not row:
            continue

        cells = [_safe_str(c) for c in row]
        non_empty = [c for c in cells if c != ""]
        if len(non_empty) < 2:
            continue

        # score: quantas strings "não numéricas" (bons headers)
        strish = 0
        for c in non_empty:
            # se parece número, não conta como header
            if re.fullmatch(r"-?\d+(\.\d+)?", c):
                continue
            strish += 1

        score = strish * 10 + len(non_empty)
        if score > best_score:
            best_score = score
            best_i = i

    return best_i


def extract_xlsx_structured(
    xlsx_bytes: bytes,
    *,
    max_cols: int = 80,
    max_rows_per_sheet: Optional[int] = None,
) -> Dict[str, Any]:
    """
    Extrai XLSX em formato estruturado + metadados.
    - Captura valores (data_only=True) e fórmula (data_only=False) quando houver.
    - Não tenta recalcular fórmula (Excel engine), mas garante rastreabilidade.
    """
    file_sha = _sha256_bytes(xlsx_bytes)

    wb_vals = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    wb_form = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=False, read_only=True)

    sheets: List[Dict[str, Any]] = []

    for ws_vals in wb_vals.worksheets:
        sheet_name = ws_vals.title
        ws_form = wb_form[sheet_name]

        # Lê tudo em memória (limita com max_rows_per_sheet se quiser)
        values_matrix: List[List[Any]] = []
        formulas_matrix: List[List[Any]] = []

        row_count = 0
        for row_vals, row_form in zip(ws_vals.iter_rows(values_only=True), ws_form.iter_rows(values_only=True)):
            row_count += 1
            if max_rows_per_sheet and row_count > max_rows_per_sheet:
                break

            row_vals = list(row_vals)[:max_cols]
            row_form = list(row_form)[:max_cols]

            values_matrix.append(row_vals)
            formulas_matrix.append(row_form)

        if not values_matrix:
            continue

        header_idx = _detect_header_row(values_matrix)
        header_raw = values_matrix[header_idx]
        header = [_norm_header(_safe_str(h)) for h in header_raw][:max_cols]

        # remove colunas vazias no final
        while header and header[-1] == "":
            header.pop()

        if not header:
            # fallback: gera colunas A,B,C...
            header = [f"COL_{i+1}" for i in range(min(len(header_raw), max_cols))]

        rows_out: List[Dict[str, Any]] = []
        schema: Dict[str, str] = {h: "null" for h in header}

        for r_i in range(header_idx + 1, len(values_matrix)):
            rv = values_matrix[r_i]
            rf = formulas_matrix[r_i]

            obj: Dict[str, Any] = {}
            any_non_empty = False

            for c_i, col in enumerate(header):
                v = rv[c_i] if c_i < len(rv) else None
                f = rf[c_i] if c_i < len(rf) else None

                # Fórmulas no openpyxl geralmente vêm como string "=..."
                formula = f if isinstance(f, str) and f.startswith("=") else None

                # Se valor é None mas tem fórmula, marca isso (sem travar)
                if v is None and formula:
                    obj[col] = None
                    obj[f"__formula__{col}"] = formula
                    obj[f"__missing__{col}"] = "formula_no_cache"
                else:
                    obj[col] = v

                if v is not None:
                    any_non_empty = True
                    t = _infer_cell_type(v)
                    if schema[col] == "null" and t != "null":
                        schema[col] = t
                    elif schema[col] != "null" and t != "null" and schema[col] != t:
                        schema[col] = "str"

            if any_non_empty:
                rows_out.append(obj)

        # hash por aba (pra versionar)
        sheet_hash_base = f"{sheet_name}|{len(header)}|{len(rows_out)}|{file_sha}"
        sheet_sha = hashlib.sha1(sheet_hash_base.encode("utf-8", errors="ignore")).hexdigest()

        sheets.append(
            {
                "sheet": sheet_name,
                "header": header,
                "rows": rows_out,
                "row_count": len(rows_out),
                "schema": schema,
                "header_row_index": header_idx,
                "sheet_sha1": sheet_sha,
            }
        )

    return {
        "file_sha256": file_sha,
        "sheets": sheets,
    }